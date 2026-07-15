from collections import OrderedDict
from datetime import date, datetime, time, timedelta
from tempfile import gettempdir
from pathlib import Path

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from database import db

router = APIRouter()


def _normalizar_fecha(fecha):
    if isinstance(fecha, str):
        return datetime.fromisoformat(fecha[:10]).date()
    return fecha


def _fecha_corta(fecha):
    return fecha.strftime("%d/%m/%Y") if hasattr(fecha, "strftime") else str(fecha)


def _parsear_fecha_periodo(periodo, fecha):
    hoy = date.today()

    if periodo not in ("diario", "semanal", "mensual"):
        raise HTTPException(status_code=400, detail="Periodo no valido")

    try:
        if not fecha:
            base = hoy
        elif periodo == "mensual" and len(fecha) == 7:
            base = datetime.strptime(f"{fecha}-01", "%Y-%m-%d").date()
        else:
            base = datetime.strptime(fecha[:10], "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Fecha no valida") from exc

    if periodo == "diario":
        inicio = base
        fin = inicio + timedelta(days=1)
        etiqueta = _fecha_corta(inicio)
    elif periodo == "semanal":
        inicio = base - timedelta(days=base.weekday())
        fin = inicio + timedelta(days=7)
        etiqueta = f"{_fecha_corta(inicio)} - {_fecha_corta(fin - timedelta(days=1))}"
    else:
        inicio = base.replace(day=1)
        if inicio.month == 12:
            fin = inicio.replace(year=inicio.year + 1, month=1)
        else:
            fin = inicio.replace(month=inicio.month + 1)
        etiqueta = inicio.strftime("%m/%Y")

    return {
        "periodo": periodo,
        "inicio": datetime.combine(inicio, time.min),
        "fin": datetime.combine(fin, time.min),
        "etiqueta": etiqueta,
    }


def _agregar_fila(sheet, row, bold=False, fill=None):
    sheet.append(row)
    if bold or fill:
        for cell in sheet[sheet.max_row]:
            if bold:
                cell.font = Font(bold=True)
            if fill:
                cell.fill = fill


def _ajustar_columnas(sheet):
    for col in sheet.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
            cell.alignment = Alignment(vertical="top")
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 42)


def _aplicar_estilo_tabla(sheet):
    header_fill = PatternFill("solid", fgColor="DDE7EE")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    _ajustar_columnas(sheet)


def _asegurar_semana(semanas, semana_inicio):
    semana_inicio = _normalizar_fecha(semana_inicio)
    key = semana_inicio.isoformat()

    if key not in semanas:
        semana_fin = semana_inicio + timedelta(days=6)
        semanas[key] = {
            "inicio": semana_inicio,
            "fin": semana_fin,
            "etiqueta": f"{_fecha_corta(semana_inicio)} - {_fecha_corta(semana_fin)}",
            "total": {
                "unidades": 0,
                "segundos": 0,
                "segundos_estandar": 0,
                "horas": 0,
                "productividad": 0,
                "eficiencia": 0,
                "operaciones": 0,
            },
            "operarios": [],
            "detalle": [],
        }

    return semanas[key]


def metricas_semanales(cursor):
    cursor.execute("""
        SELECT
            DATE_TRUNC('week', r.inicio::timestamp)::date AS semana_inicio,
            op.nombre,
            COALESCE(SUM(r.cantidad), 0) AS unidades,
            COALESCE(SUM(EXTRACT(EPOCH FROM (r.fin::timestamp - r.inicio::timestamp))), 0) AS segundos,
            COUNT(r.id) AS operaciones,
            COALESCE(SUM(CASE WHEN e.unidades_por_hora > 0 THEN (r.cantidad::numeric / e.unidades_por_hora) * 3600 ELSE 0 END), 0) AS segundos_estandar
        FROM registros_produccion r
        JOIN operarios op ON op.id = r.operario_id
        JOIN actividades a ON a.id = r.actividad_id
        LEFT JOIN estandares_actividad e ON e.actividad_id = a.id
        GROUP BY DATE_TRUNC('week', r.inicio::timestamp)::date, op.id, op.nombre
        ORDER BY semana_inicio DESC, unidades DESC, op.nombre
    """)

    resumen = cursor.fetchall()
    semanas = OrderedDict()
    operarios_por_semana = {}

    for semana_inicio, nombre, unidades, segundos, operaciones, segundos_estandar in resumen:
        semana = _asegurar_semana(semanas, semana_inicio)
        segundos = float(segundos or 0)
        segundos_estandar = float(segundos_estandar or 0)
        unidades = int(unidades or 0)
        operaciones = int(operaciones or 0)
        horas = round(segundos / 3600, 2) if segundos else 0
        horas_estandar = round(segundos_estandar / 3600, 2) if segundos_estandar else 0
        productividad = round(unidades / horas, 2) if horas else 0
        eficiencia = round((horas_estandar / horas) * 100, 1) if horas else 0
        utilizacion = round((horas / 48) * 100, 1)

        operario_data = {
            "nombre": nombre,
            "unidades": unidades,
            "horas": horas,
            "eficiencia": eficiencia,
            "utilizacion": utilizacion,
            "productividad": productividad,
            "operaciones": operaciones,
            "registros": [],
        }

        semana["operarios"].append(operario_data)
        operarios_por_semana[(semana["inicio"], nombre)] = operario_data

        semana["total"]["unidades"] += unidades
        semana["total"]["segundos"] += segundos
        semana["total"]["segundos_estandar"] += segundos_estandar
        semana["total"]["operaciones"] += operaciones

    cursor.execute("""
        SELECT
            DATE_TRUNC('week', r.inicio::timestamp)::date AS semana_inicio,
            op.nombre,
            a.nombre,
            r.cantidad,
            r.inicio,
            r.fin,
            ROUND((EXTRACT(EPOCH FROM (r.fin::timestamp - r.inicio::timestamp)) / 3600.0)::numeric, 2) AS horas,
            CASE WHEN e.unidades_por_hora > 0 THEN ROUND((r.cantidad::numeric / e.unidades_por_hora), 2) ELSE 0 END AS horas_estandar
        FROM registros_produccion r
        JOIN operarios op ON op.id = r.operario_id
        JOIN actividades a ON a.id = r.actividad_id
        LEFT JOIN estandares_actividad e ON e.actividad_id = a.id
        ORDER BY semana_inicio DESC, r.inicio::timestamp DESC, r.id DESC
    """)

    detalle = cursor.fetchall()

    for semana_inicio, operario, actividad, cantidad, inicio, fin, horas, horas_estandar in detalle:
        semana = _asegurar_semana(semanas, semana_inicio)
        horas = float(horas or 0)
        horas_estandar = float(horas_estandar or 0)
        
        registro = {
            "operario": operario,
            "actividad": actividad,
            "cantidad": cantidad,
            "inicio": inicio,
            "fin": fin,
            "horas": horas,
            "eficiencia": round((horas_estandar / horas) * 100, 1) if horas > 0 else 0
        }

        semana["detalle"].append(registro)

        operario_data = operarios_por_semana.get((semana["inicio"], operario))
        if operario_data:
            operario_data["registros"].append(registro)

    for semana in semanas.values():
        segundos = semana["total"]["segundos"]
        segundos_estandar = semana["total"]["segundos_estandar"]
        horas = round(segundos / 3600, 2) if segundos else 0
        horas_estandar = round(segundos_estandar / 3600, 2) if segundos_estandar else 0
        
        semana["total"]["horas"] = horas
        semana["total"]["eficiencia"] = round((horas_estandar / horas) * 100, 1) if horas else 0
        semana["total"]["productividad"] = (
            round(semana["total"]["unidades"] / horas, 2) if horas else 0
        )

    return list(semanas.values())


def _horas_productividad(segundos, unidades):
    segundos = float(segundos or 0)
    unidades = int(unidades or 0)
    horas = round(segundos / 3600, 2) if segundos else 0
    productividad = round(unidades / horas, 2) if horas else 0
    return horas, productividad


def _crear_excel_metricas(cursor, filtro):
    inicio = filtro["inicio"]
    fin = filtro["fin"]
    params = (inicio, fin)

    cursor.execute("""
        SELECT
            op.nombre,
            COALESCE(SUM(r.cantidad), 0) AS unidades,
            COALESCE(SUM(EXTRACT(EPOCH FROM (r.fin::timestamp - r.inicio::timestamp))), 0) AS segundos,
            COUNT(r.id) AS operaciones
        FROM registros_produccion r
        JOIN operarios op ON op.id = r.operario_id
        WHERE r.inicio::timestamp >= %s
        AND r.inicio::timestamp < %s
        GROUP BY op.id, op.nombre
        ORDER BY unidades DESC, op.nombre
    """, params)
    resumen_operarios = cursor.fetchall()

    cursor.execute("""
        SELECT
            r.inicio::date AS fecha,
            op.nombre,
            COALESCE(SUM(r.cantidad), 0) AS unidades,
            COALESCE(SUM(EXTRACT(EPOCH FROM (r.fin::timestamp - r.inicio::timestamp))), 0) AS segundos,
            COUNT(r.id) AS operaciones
        FROM registros_produccion r
        JOIN operarios op ON op.id = r.operario_id
        WHERE r.inicio::timestamp >= %s
        AND r.inicio::timestamp < %s
        GROUP BY r.inicio::date, op.id, op.nombre
        ORDER BY fecha, op.nombre
    """, params)
    resumen_diario = cursor.fetchall()

    cursor.execute("""
        SELECT
            a.nombre,
            op.nombre,
            COALESCE(SUM(r.cantidad), 0) AS unidades,
            COALESCE(SUM(EXTRACT(EPOCH FROM (r.fin::timestamp - r.inicio::timestamp))), 0) AS segundos,
            COUNT(r.id) AS operaciones
        FROM registros_produccion r
        JOIN operarios op ON op.id = r.operario_id
        JOIN actividades a ON a.id = r.actividad_id
        WHERE r.inicio::timestamp >= %s
        AND r.inicio::timestamp < %s
        GROUP BY a.id, a.nombre, op.id, op.nombre
        ORDER BY a.nombre, unidades DESC, op.nombre
    """, params)
    resumen_actividad = cursor.fetchall()

    cursor.execute("""
        SELECT
            r.id,
            r.inicio::date AS fecha,
            op.nombre AS operario,
            COALESCE(m.nombre, '') AS maquina,
            r.orden_id,
            COALESCE(p.nombre, '') AS proceso,
            a.nombre AS actividad,
            r.cantidad,
            r.inicio,
            r.fin,
            ROUND((EXTRACT(EPOCH FROM (r.fin::timestamp - r.inicio::timestamp)) / 3600.0)::numeric, 2) AS horas,
            CASE WHEN e.unidades_por_hora > 0 THEN ROUND((r.cantidad::numeric / e.unidades_por_hora), 2) ELSE 0 END AS horas_estandar
        FROM registros_produccion r
        JOIN operarios op ON op.id = r.operario_id
        JOIN actividades a ON a.id = r.actividad_id
        LEFT JOIN estandares_actividad e ON e.actividad_id = a.id
        LEFT JOIN procesos p ON p.id = a.proceso_id
        LEFT JOIN ordenes o ON o.id = r.orden_id
        LEFT JOIN maquinas m ON m.id = o.maquina_id
        WHERE r.inicio::timestamp >= %s
        AND r.inicio::timestamp < %s
        ORDER BY r.inicio::timestamp, op.nombre, r.id
    """, params)
    registros = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    total_unidades = sum(int(row[1] or 0) for row in resumen_operarios)
    total_segundos = sum(float(row[2] or 0) for row in resumen_operarios)
    total_operaciones = sum(int(row[3] or 0) for row in resumen_operarios)
    total_horas, total_productividad = _horas_productividad(total_segundos, total_unidades)

    titulo = f"Metricas {filtro['periodo']} - {filtro['etiqueta']}"
    ws.append([titulo])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Desde", inicio.strftime("%Y-%m-%d %H:%M")])
    ws.append(["Hasta", (fin - timedelta(seconds=1)).strftime("%Y-%m-%d %H:%M")])
    ws.append([])
    ws.append(["Total unidades", "Total horas", "Unid/Hora", "Operaciones"])
    ws.append([total_unidades, total_horas, total_productividad, total_operaciones])
    ws.append([])
    _agregar_fila(ws, ["Operario", "Unidades", "Horas", "Unid/Hora", "Operaciones"], bold=True)

    for nombre, unidades, segundos, operaciones in resumen_operarios:
        horas, productividad = _horas_productividad(segundos, unidades)
        ws.append([nombre, int(unidades or 0), horas, productividad, int(operaciones or 0)])

    ws_por_dia = wb.create_sheet("Por dia")
    _agregar_fila(ws_por_dia, ["Fecha", "Operario", "Unidades", "Horas", "Unid/Hora", "Operaciones"], bold=True)
    for fecha, operario, unidades, segundos, operaciones in resumen_diario:
        horas, productividad = _horas_productividad(segundos, unidades)
        ws_por_dia.append([fecha, operario, int(unidades or 0), horas, productividad, int(operaciones or 0)])

    ws_actividad = wb.create_sheet("Por actividad")
    _agregar_fila(ws_actividad, ["Actividad", "Operario", "Unidades", "Horas", "Unid/Hora", "Operaciones"], bold=True)
    for actividad, operario, unidades, segundos, operaciones in resumen_actividad:
        horas, productividad = _horas_productividad(segundos, unidades)
        ws_actividad.append([actividad, operario, int(unidades or 0), horas, productividad, int(operaciones or 0)])

    ws_detalle = wb.create_sheet("Registros")
    _agregar_fila(
        ws_detalle,
        ["ID", "Fecha", "Operario", "Maquina", "Orden", "Proceso", "Actividad", "Cantidad", "Inicio", "Fin", "Horas", "Unid/Hora", "Eficiencia %"],
        bold=True,
    )
    for registro_id, fecha, operario, maquina, orden_id, proceso, actividad, cantidad, inicio_reg, fin_reg, horas, horas_estandar in registros:
        horas = float(horas or 0)
        horas_estandar = float(horas_estandar or 0)
        cantidad = int(cantidad or 0)
        productividad = round(cantidad / horas, 2) if horas else 0
        eficiencia = round((horas_estandar / horas) * 100, 1) if horas > 0 else 0
        ws_detalle.append([
            registro_id,
            fecha,
            operario,
            maquina,
            orden_id,
            proceso,
            actividad,
            cantidad,
            inicio_reg,
            fin_reg,
            horas,
            productividad,
            eficiencia,
        ])

    for sheet in wb.worksheets:
        _ajustar_columnas(sheet)

    header_fill = PatternFill("solid", fgColor="DDE7EE")
    for sheet in (ws_por_dia, ws_actividad, ws_detalle):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"

    for cell in ws[8]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    nombre_base = f"metricas_{filtro['periodo']}_{inicio.strftime('%Y%m%d')}_{(fin - timedelta(days=1)).strftime('%Y%m%d')}"
    nombre_archivo = f"{nombre_base}.xlsx"
    ruta = Path(gettempdir()) / f"{nombre_base}_{datetime.now().strftime('%H%M%S%f')}.xlsx"
    wb.save(ruta)

    return ruta, nombre_archivo


@router.get("/metricas_operarios", response_class=HTMLResponse)
def metricas_operarios(request: Request):
    if not request.session.get("role") == "admin":
        return RedirectResponse("/admin", 303)

    conn = db()
    c = conn.cursor()
    semanas = metricas_semanales(c)
    conn.close()

    return request.app.state.templates.TemplateResponse(
        request=request,
        name="Metricas.html",
        context={
            "request": request,
            "semanas": semanas,
        },
    )


@router.get("/metricas", response_class=HTMLResponse)
def metricas(request: Request):
    if not request.session.get("role") == "admin":
        return RedirectResponse("/admin", 303)

    conn = db()
    c = conn.cursor()
    semanas = metricas_semanales(c)
    conn.close()

    return request.app.state.templates.TemplateResponse(
        request=request,
        name="Metricas.html",
        context={
            "request": request,
            "semanas": semanas,
        },
    )


@router.get("/metricas/exportar_excel")
@router.get("/exportar_excel")
def exportar_metricas_excel(request: Request, periodo: str = "semanal", fecha: str = None):
    if not request.session.get("role") == "admin":
        return RedirectResponse("/admin", 303)

    filtro = _parsear_fecha_periodo(periodo, fecha)

    conn = db()
    c = conn.cursor()
    ruta, nombre_archivo = _crear_excel_metricas(c, filtro)
    conn.close()

    return FileResponse(
        path=str(ruta),
        filename=nombre_archivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/kpi", response_class=HTMLResponse)
def kpi(request: Request):
    if not request.session.get("role") == "admin":
        return RedirectResponse("/admin", 303)

    conn = db()
    c = conn.cursor()

    c.execute("""
    SELECT o.nombre,
           COALESCE(SUM(r.cantidad),0)
    FROM registros_produccion r
    JOIN operarios o ON o.id=r.operario_id
    GROUP BY o.id
    """)
    por_operario = c.fetchall()

    c.execute("""
    SELECT o.nombre,
           COALESCE(SUM(EXTRACT(EPOCH FROM (r.fin::timestamp - r.inicio::timestamp)) / 60.0), 0)
    FROM registros_produccion r
    JOIN operarios o ON o.id=r.operario_id
    GROUP BY o.id
    """)
    minutos = c.fetchall()

    c.execute("""
    SELECT substr(inicio,1,10),
           SUM(cantidad)
    FROM registros_produccion
    GROUP BY substr(inicio,1,10)
    ORDER BY substr(inicio,1,10)
    """)
    diario = c.fetchall()

    conn.close()

    return request.app.state.templates.TemplateResponse(
        request=request,
        name="kpi.html",
        context={
            "request": request,
            "por_operario": por_operario,
            "minutos": minutos,
            "diario": diario,
        },
    )
